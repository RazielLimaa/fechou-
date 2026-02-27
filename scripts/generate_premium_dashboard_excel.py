#!/usr/bin/env python3
import json
import sys
import datetime as dt
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import LineChart, BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

BRAND_ORANGE = "FFFF6600"
BRAND_DARK = "FF0B1220"
BRAND_DARK2 = "FF111827"
BRAND_GRAY = "FFF3F4F6"
BRAND_TEXT = "FF111827"
BRAND_WHITE = "FFFFFFFF"


def col_width(ws, col, width):
    ws.column_dimensions[col].width = width


def row_height(ws, row, height):
    ws.row_dimensions[row].height = height


def thin_border(color="FF1F2937"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def set_area(ws, r1, c1, r2, c2, fill=None, border=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            if fill:
                cell.fill = fill
            if border:
                cell.border = border


def merge(ws, r1, c1, r2, c2, value=None, font=None, fill=None, align=None, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    set_area(ws, r1, c1, r2, c2, fill=fill, border=border)
    return cell


def kpi_card(ws, top, left, title, formula_cell, icon="●"):
    r, c = top, left
    set_area(ws, r, c, r + 6, c + 3, fill=PatternFill("solid", fgColor=BRAND_WHITE), border=thin_border("FFE5E7EB"))
    merge(ws, r, c, r, c + 3, value=f"{icon}  {title}",
          font=Font(bold=True, color=BRAND_TEXT, size=11),
          fill=PatternFill("solid", fgColor=BRAND_GRAY),
          align=Alignment(horizontal="left", vertical="center"),
          border=thin_border("FFE5E7EB"))
    merge(ws, r + 1, c, r + 3, c + 3, value=f"={formula_cell}",
          font=Font(bold=True, color=BRAND_ORANGE, size=20),
          fill=PatternFill("solid", fgColor=BRAND_WHITE),
          align=Alignment(horizontal="left", vertical="center"),
          border=thin_border("FFE5E7EB"))
    merge(ws, r + 4, c, r + 6, c + 3, value="Auto (aba Dados).",
          font=Font(color="FF6B7280", size=9),
          fill=PatternFill("solid", fgColor=BRAND_WHITE),
          align=Alignment(horizontal="left", vertical="top", wrap_text=True),
          border=thin_border("FFE5E7EB"))
    row_height(ws, r, 20)
    for rr in range(r + 1, r + 4):
        row_height(ws, rr, 22)
    for rr in range(r + 4, r + 7):
        row_height(ws, rr, 16)


def parse_value(raw):
    if raw is None:
        return 0.0
    text = str(raw).strip()
    if not text:
        return 0.0
    return float(text.replace('.', '').replace(',', '.'))


def parse_date(raw):
    if isinstance(raw, dt.date):
        return raw
    value = str(raw)
    try:
        return dt.datetime.fromisoformat(value.replace('Z', '+00:00')).date()
    except Exception:
        return dt.date.today()


def main():
    if len(sys.argv) < 2:
        print("usage: generate_premium_dashboard_excel.py <output.xlsx>", file=sys.stderr)
        sys.exit(2)

    out_path = sys.argv[1]
    payload = json.loads(sys.stdin.read() or "{}")
    proposals = payload.get("proposals", [])

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_data = wb.create_sheet("Dados")
    ws_clients = wb.create_sheet("Clientes")
    ws_cfg = wb.create_sheet("Config")

    ws_cfg.sheet_view.showGridLines = False
    merge(ws_cfg, 1, 1, 1, 4, "Configurações — Fechou!", Font(bold=True, size=14, color=BRAND_TEXT))
    ws_cfg["A3"], ws_cfg["B3"] = "Moeda", "BRL"
    ws_cfg["A4"], ws_cfg["B4"] = "Período (meses)", 12
    ws_cfg["A5"], ws_cfg["B5"] = "Ano base", dt.date.today().year
    for r in range(3, 6):
        ws_cfg[f"A{r}"].font = Font(bold=True, color=BRAND_TEXT)
        ws_cfg[f"B{r}"].font = Font(color="FF2563EB")
    col_width(ws_cfg, "A", 22)
    col_width(ws_cfg, "B", 14)

    headers = ["Data", "Cliente", "Proposta", "Status", "Canal", "Valor (R$)", "Pago?", "Vendedor", "Categoria", "UF"]
    ws_data.append(headers)

    channels = ["Pix", "Cartão", "Boleto"]
    sellers = ["Razi", "Ana", "João", "Bea"]
    cats = ["Serviço", "Produto", "Mensalidade", "Setup"]
    ufs = ["SP", "RJ", "MG", "PR"]

    rows = []
    for p in proposals:
        pid = int(p.get("id", 0) or 0)
        status = str(p.get("status", "pendente"))
        d = parse_date(p.get("createdAt"))
        val = round(parse_value(p.get("value")), 2)
        rows.append([
            d,
            str(p.get("clientName", "")),
            str(p.get("title", f"Contrato #{pid}")),
            status,
            channels[pid % len(channels)],
            val,
            "Sim" if status == "vendida" else "Não",
            sellers[(pid + 1) % len(sellers)],
            cats[(pid + 2) % len(cats)],
            ufs[(pid + 3) % len(ufs)],
        ])

    rows.sort(key=lambda x: x[0])
    for r in rows:
        ws_data.append(r)

    ws_data.freeze_panes = "A2"
    ws_data.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor=BRAND_DARK2)
    header_font = Font(bold=True, color=BRAND_WHITE)
    for col in range(1, len(headers) + 1):
        cell = ws_data.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border(BRAND_DARK2)

    width_map = [12, 20, 18, 12, 10, 14, 8, 12, 14, 6]
    for i, w in enumerate(width_map, start=1):
        col_width(ws_data, get_column_letter(i), w)

    max_row = ws_data.max_row
    for r in range(2, max_row + 1):
        ws_data.cell(r, 1).number_format = "yyyy-mm-dd"
        ws_data.cell(r, 6).number_format = '"R$" #,##0.00'
        for c in range(1, len(headers) + 1):
            cell = ws_data.cell(r, c)
            cell.alignment = Alignment(vertical="center", horizontal="left")
            cell.border = thin_border("FFE5E7EB")

    if max_row >= 2:
        ws_data.conditional_formatting.add(
            f"D2:D{max_row}",
            CellIsRule(operator="equal", formula=['"vendida"'], fill=PatternFill("solid", fgColor="FFD1FAE5"))
        )
        ws_data.conditional_formatting.add(
            f"D2:D{max_row}",
            CellIsRule(operator="equal", formula=['"pendente"'], fill=PatternFill("solid", fgColor="FFFEF3C7"))
        )
        ws_data.conditional_formatting.add(
            f"D2:D{max_row}",
            CellIsRule(operator="equal", formula=['"cancelada"'], fill=PatternFill("solid", fgColor="FFFEE2E2"))
        )

    ws_clients.sheet_view.showGridLines = False
    merge(ws_clients, 1, 1, 1, 7, "Clientes — Ranking e Métricas", Font(bold=True, size=16, color=BRAND_TEXT))
    merge(ws_clients, 2, 1, 2, 7, "Baseado na aba Dados. Ideal para priorizar follow-up e ver receita por cliente.",
          Font(color="FF6B7280", size=10), align=Alignment(horizontal="left"))

    client_headers = ["Cliente", "Receita (R$)", "Vendas", "Pendentes", "Canceladas", "Ticket Médio (R$)", "Última Venda"]
    ws_clients.append([])
    ws_clients.append(client_headers)

    for i in range(1, 8):
        cell = ws_clients.cell(4, i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border(BRAND_DARK2)

    names = sorted({r[1] for r in rows})
    start = 5
    for idx, name in enumerate(names):
        r = start + idx
        ws_clients.cell(r, 1, name)
        ws_clients.cell(r, 2, f'=SUMIFS(Dados!$F:$F,Dados!$B:$B,$A{r},Dados!$D:$D,"vendida")')
        ws_clients.cell(r, 3, f'=COUNTIFS(Dados!$B:$B,$A{r},Dados!$D:$D,"vendida")')
        ws_clients.cell(r, 4, f'=COUNTIFS(Dados!$B:$B,$A{r},Dados!$D:$D,"pendente")')
        ws_clients.cell(r, 5, f'=COUNTIFS(Dados!$B:$B,$A{r},Dados!$D:$D,"cancelada")')
        ws_clients.cell(r, 6, f'=IFERROR($B{r}/$C{r},0)')
        ws_clients.cell(r, 7, f'=IFERROR(MAXIFS(Dados!$A:$A,Dados!$B:$B,$A{r},Dados!$D:$D,"vendida"),"")')
        ws_clients.cell(r, 2).number_format = '"R$" #,##0.00'
        ws_clients.cell(r, 6).number_format = '"R$" #,##0.00'
        ws_clients.cell(r, 7).number_format = "yyyy-mm-dd"
        for c in range(1, 8):
            cell = ws_clients.cell(r, c)
            cell.border = thin_border("FFE5E7EB")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    total_r = start + len(names)
    ws_clients.cell(total_r, 1, "TOTAL").font = Font(bold=True)
    ws_clients.cell(total_r, 2, f"=SUM(B{start}:B{total_r-1})")
    ws_clients.cell(total_r, 3, f"=SUM(C{start}:C{total_r-1})")
    ws_clients.cell(total_r, 4, f"=SUM(D{start}:D{total_r-1})")
    ws_clients.cell(total_r, 5, f"=SUM(E{start}:E{total_r-1})")
    ws_clients.cell(total_r, 6, f"=IFERROR(B{total_r}/C{total_r},0)")
    ws_clients.cell(total_r, 2).number_format = '"R$" #,##0.00'
    ws_clients.cell(total_r, 6).number_format = '"R$" #,##0.00'
    for c in range(1, 8):
        cell = ws_clients.cell(total_r, c)
        cell.fill = PatternFill("solid", fgColor=BRAND_GRAY)
        cell.border = thin_border("FFE5E7EB")

    for col, w in zip(["A", "B", "C", "D", "E", "F", "G"], [22, 16, 10, 12, 14, 16, 14]):
        col_width(ws_clients, col, w)
    ws_clients.freeze_panes = "A5"

    ws_dash.sheet_view.showGridLines = False
    for col, w in zip(list("ABCDEFGHIJKLMNOPQRST"), [2,16,16,16,2,16,16,16,2,16,16,16,2,16,16,16,2,16,16,16]):
        col_width(ws_dash, col, w)

    set_area(ws_dash, 1, 1, 4, 20, fill=PatternFill("solid", fgColor=BRAND_DARK), border=thin_border(BRAND_DARK))
    merge(ws_dash, 2, 2, 3, 3, "F!", font=Font(bold=True, size=40, color=BRAND_ORANGE),
          fill=PatternFill("solid", fgColor=BRAND_DARK), align=Alignment(horizontal="left", vertical="center"))
    merge(ws_dash, 2, 4, 2, 19, "Relatório Executivo — Fechou!", font=Font(bold=True, size=20, color=BRAND_WHITE),
          fill=PatternFill("solid", fgColor=BRAND_DARK), align=Alignment(horizontal="left", vertical="center"))
    merge(ws_dash, 3, 4, 3, 19, "KPIs • gráficos • ranking de clientes (modelo premium)",
          font=Font(size=10, color="FFCBD5E1"), fill=PatternFill("solid", fgColor=BRAND_DARK),
          align=Alignment(horizontal="left", vertical="center"))

    ws_dash["S6"] = '=SUMIFS(Dados!$F:$F,Dados!$D:$D,"vendida")'
    ws_dash["S7"] = '=COUNTIF(Dados!$D:$D,"vendida")'
    ws_dash["S8"] = '=COUNTIF(Dados!$D:$D,"pendente")'
    ws_dash["S9"] = '=COUNTIF(Dados!$D:$D,"cancelada")'
    ws_dash["S10"] = '=IFERROR(S6/S7,0)'
    ws_dash["S11"] = '=IFERROR(S7/(S7+S8+S9),0)'
    ws_dash["S6"].number_format = '"R$" #,##0.00'
    ws_dash["S10"].number_format = '"R$" #,##0.00'
    ws_dash["S11"].number_format = "0.0%"

    kpi_card(ws_dash, 6, 2, "Receita total", "Dashboard!$S$6", icon="💰")
    kpi_card(ws_dash, 6, 7, "Vendas (qtd)", "Dashboard!$S$7", icon="✅")
    kpi_card(ws_dash, 6, 12, "Pendentes (qtd)", "Dashboard!$S$8", icon="⏳")
    kpi_card(ws_dash, 14, 2, "Ticket médio", "Dashboard!$S$10", icon="📈")
    kpi_card(ws_dash, 14, 7, "Conversão", "Dashboard!$S$11", icon="🎯")
    kpi_card(ws_dash, 14, 12, "Canceladas (qtd)", "Dashboard!$S$9", icon="🧯")

    merge(ws_dash, 22, 2, 22, 9, "Tendência — últimos 12 meses", font=Font(bold=True, size=12, color=BRAND_TEXT))
    series_headers = ["Mês", "Receita (R$)", "Vendidas", "Pendentes", "Canceladas", "Pix (R$)"]
    for i, h in enumerate(series_headers, start=2):
        c = ws_dash.cell(23, i)
        c.value = h
        c.fill = PatternFill("solid", fgColor=BRAND_GRAY)
        c.font = Font(bold=True, color=BRAND_TEXT)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border("FFE5E7EB")

    a_rng = f'Dados!$A$2:$A${max_row}'
    d_rng = f'Dados!$D$2:$D${max_row}'
    f_rng = f'Dados!$F$2:$F${max_row}'
    canal_rng = f'Dados!$E$2:$E${max_row}'

    base_row = 24
    for i in range(12):
        r = base_row + i
        ws_dash.cell(r, 2).value = f"=TEXT(EDATE(TODAY(),-{11-i}),\"yyyy-mm\")"
        ws_dash.cell(r, 3).value = f'=SUMPRODUCT((TEXT({a_rng},"yyyy-mm")=B{r})*({d_rng}="vendida")*({f_rng}))'
        ws_dash.cell(r, 4).value = f'=SUMPRODUCT((TEXT({a_rng},"yyyy-mm")=B{r})*({d_rng}="vendida"))'
        ws_dash.cell(r, 5).value = f'=SUMPRODUCT((TEXT({a_rng},"yyyy-mm")=B{r})*({d_rng}="pendente"))'
        ws_dash.cell(r, 6).value = f'=SUMPRODUCT((TEXT({a_rng},"yyyy-mm")=B{r})*({d_rng}="cancelada"))'
        ws_dash.cell(r, 7).value = f'=SUMPRODUCT((TEXT({a_rng},"yyyy-mm")=B{r})*({d_rng}="vendida")*({canal_rng}="Pix")*({f_rng}))'
        ws_dash.cell(r, 3).number_format = '"R$" #,##0.00'
        ws_dash.cell(r, 7).number_format = '"R$" #,##0.00'
        for c in range(2, 8):
            cell = ws_dash.cell(r, c)
            cell.border = thin_border("FFE5E7EB")
            cell.alignment = Alignment(horizontal="left")

    cats_ref = Reference(ws_dash, min_col=2, min_row=24, max_row=35)

    line = LineChart()
    line.title = "Receita por mês (vendidas)"
    line.y_axis.title = "R$"
    line.x_axis.title = "Mês"
    line.add_data(Reference(ws_dash, min_col=3, min_row=23, max_row=35), titles_from_data=True)
    line.set_categories(cats_ref)
    line.height = 10
    line.width = 24
    ws_dash.add_chart(line, "I23")

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "Volume por status (qtd)"
    bar.y_axis.title = "Qtd"
    bar.x_axis.title = "Mês"
    bar.add_data(Reference(ws_dash, min_col=4, min_row=23, max_col=6, max_row=35), titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.height = 10
    bar.width = 24
    ws_dash.add_chart(bar, "I36")

    merge(ws_dash, 36, 2, 36, 8, "Distribuição geral", font=Font(bold=True, size=12, color=BRAND_TEXT))
    ws_dash["B37"], ws_dash["C37"] = "Status", "Qtd"
    ws_dash["B38"], ws_dash["C38"] = "Vendida", "=Dashboard!$S$7"
    ws_dash["B39"], ws_dash["C39"] = "Pendente", "=Dashboard!$S$8"
    ws_dash["B40"], ws_dash["C40"] = "Cancelada", "=Dashboard!$S$9"
    for r in range(37, 41):
        for c in [2, 3]:
            cell = ws_dash.cell(r, c)
            cell.border = thin_border("FFE5E7EB")
            cell.alignment = Alignment(horizontal="left")
    ws_dash["B37"].fill = PatternFill("solid", fgColor=BRAND_GRAY)
    ws_dash["C37"].fill = PatternFill("solid", fgColor=BRAND_GRAY)
    ws_dash["B37"].font = Font(bold=True, color=BRAND_TEXT)
    ws_dash["C37"].font = Font(bold=True, color=BRAND_TEXT)

    dough = DoughnutChart()
    dough.title = "Status (geral)"
    dough.add_data(Reference(ws_dash, min_col=3, min_row=37, max_row=40), titles_from_data=True)
    dough.set_categories(Reference(ws_dash, min_col=2, min_row=38, max_row=40))
    dough.dataLabels = DataLabelList()
    dough.dataLabels.showPercent = True
    dough.height = 9
    dough.width = 12
    ws_dash.add_chart(dough, "E37")

    merge(ws_dash, 42, 9, 42, 16, "Top 10 clientes por receita", font=Font(bold=True, size=12, color=BRAND_TEXT))
    ws_dash["I43"], ws_dash["J43"], ws_dash["K43"] = "Cliente", "Receita (R$)", "Vendas"
    for cell in [ws_dash["I43"], ws_dash["J43"], ws_dash["K43"]]:
        cell.fill = PatternFill("solid", fgColor=BRAND_GRAY)
        cell.font = Font(bold=True, color=BRAND_TEXT)
        cell.border = thin_border("FFE5E7EB")
        cell.alignment = Alignment(horizontal="center")

    sorted_clients = sorted(names, key=lambda n: sum((r[5] for r in rows if r[1] == n and r[3] == "vendida")), reverse=True)
    for i in range(10):
        rr = 44 + i
        if i < len(sorted_clients):
            name = sorted_clients[i]
            ws_dash[f"I{rr}"] = name
            ws_dash[f"J{rr}"] = f'=SUMIFS(Dados!$F:$F,Dados!$B:$B,I{rr},Dados!$D:$D,"vendida")'
            ws_dash[f"K{rr}"] = f'=COUNTIFS(Dados!$B:$B,I{rr},Dados!$D:$D,"vendida")'
            ws_dash[f"J{rr}"].number_format = '"R$" #,##0.00'
        for col in ["I", "J", "K"]:
            ws_dash[f"{col}{rr}"].border = thin_border("FFE5E7EB")
            ws_dash[f"{col}{rr}"].alignment = Alignment(horizontal="left")

    ws_dash.freeze_panes = "A5"
    wb.save(out_path)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        sys.exit(1)
